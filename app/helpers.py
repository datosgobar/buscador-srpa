import csv
from .models import Question, Topic, SubTopic, Author, Report
import math
from flask import request, url_for, g
from sqlalchemy import func
from textar import TextClassifier
from datetime import datetime
from openpyxl import load_workbook


class SpreadSheetReader:

    def __init__(self):
        pass

    @classmethod
    def first_read(cls, filename):
        extension = filename.split('.')[-1]
        file_path = 'app/uploads/' + filename
        if extension == 'csv':
            spreadsheet = cls.read_csv(file_path)
        elif extension == 'xlsx':
            spreadsheet = cls.read_xlsx(file_path)
        else:
            raise Exception('Formato no soportado')

        summary = {'best_row': []}
        data = []
        for i, row in spreadsheet:
            if i == 0:
                summary['first_row'] = row
                data = [[] for col in row]
                continue
            for colnum in range(len(data)):
                data[colnum].append(str(row[colnum]))
            summary['best_row'] = cls._best_row(summary['best_row'], row)
        summary['datatypes'] = cls._guess_datatypes(data)
        return summary

    @classmethod
    def read_csv(cls, csv_path):
        with open(csv_path, 'r', encoding='utf-8') as csvfile:
            dialect = csv.Sniffer().sniff(csvfile.read(), delimiters=',')
            csvfile.seek(0)
            reader = csv.reader(csvfile, dialect)
            for i, row in enumerate(reader):
                yield (i, [str(cell) for cell in row])

    @classmethod
    def read_xlsx(cls, xlsx_file_path):
        wb = load_workbook(xlsx_file_path, read_only=True)
        first_sheet = wb[wb.sheetnames[0]]

        def cell_value(cell):
            if cell.is_date:
                return cell.value.date()
            return str(cell.value or '').strip()

        for i, row in enumerate(first_sheet):
            yield (i, [cell_value(cell) for cell in row])

    @staticmethod
    def _best_row(first_row, second_row):
        first_row = [str(value) for value in first_row]
        second_row = [str(value) for value in second_row]

        def columns_with_values(row):
            return sum([1 for field in row if len(field.strip()) > 0])

        def average_column_length(row):
            return sum([len(field.strip()) for field in row]) / float(len(row))

        first_columns_with_values = columns_with_values(first_row)
        second_columns_with_values = columns_with_values(second_row)
        if first_columns_with_values > second_columns_with_values:
            return first_row
        elif second_columns_with_values > first_columns_with_values:
            return second_row
        else:
            first_average = average_column_length(first_row)
            second_average = average_column_length(second_row)
            if first_average > second_average:
                return first_row
            return second_row

    @staticmethod
    def _guess_datatypes(data):
        """ Recibe una lista de listas [col1, col2, col3..] e intenta adivinar
        el tipo de dato en cada una. Por ahora solo hay 4 tipos:
        `Numero` `Texto` `Categoria` `Otro`
        """
        data_props = [{} for col in data]
        for i, col in enumerate(data):
            data_is_empty = map(lambda x: len(x) == 0, col)
            if all(data_is_empty):
                data_props[i]['warnings'] = '¡Columna Vacia!'
                data_props[i]['types'] = ['']
                continue
            else:
                data_props[i]['warnings'] = ''
            if any(data_is_empty):
                data_props[i]['empty_status'] = 'Contiene Vacios'
            else:
                data_props[i]['empty_status'] = ''
            non_empty = list(filter(lambda x: len(x) > 0, col))
            data_props[i]['types'] = []
            types = data_props[i]['types']
            if all(map(lambda x: x.isdigit(), non_empty)):
                types.append('Numeros')
            else:
                types.append('Letras')
            if len(set(non_empty)) < (len(non_empty) * 0.5):
                types.append('Categoria')
            elif (sum(map(len, non_empty)) / len(non_empty) > 100 and
                    'Letras' in types):
                types[types.index('Letras')] = 'Texto'
            else:
                types.append('Otro')
        return data_props


class Searcher:

    def __init__(self):
        self.text_classifier = None
        self.restart_text_classifier()
        self.per_page = 10

    def restart_text_classifier(self):
        all_questions = Question.query.all()
        if len(all_questions) > 0:
            qids = ['q' + str(q.id) for q in all_questions
                    if q.body is not None]
            rids = ['r' + str(q.id) for q in all_questions
                    if q.answer is not None]
            q_texts = [q.body + q.context for q in all_questions
                       if q.body is not None]
            r_texts = [q.answer for q in all_questions
                       if q.answer is not None]
            try:
                self.text_classifier = TextClassifier(q_texts + r_texts,
                                                      qids + rids)
                self.restart_suggesters(all_questions)
            except Exception as e:
                print(e)

    def restart_suggesters(self, questions):
        ids = ['q' + str(q.id) for q in questions if q.topic is not None]
        topic_ids = [str(q.topic.id) for q in questions if q.topic is not None]
        self.text_classifier.make_classifier("topics", ids, topic_ids)
        all_topics = Topic.query.all()
        for topic in all_topics:
            if len(topic.name) < 2:
                continue
            questions_with_topic = Question.query.filter_by(topic=topic).all()
            if len(set(questions_with_topic)) > 2:
                questions_with_topic_ids = [str(q.id) for q in questions_with_topic if q.subtopic is not None]
                subtopic_ids = [str(q.subtopic.id) for q in questions_with_topic if q.subtopic is not None]
                classifier_name = str(topic.id) + "_subtopics"
                self.text_classifier.make_classifier(classifier_name, questions_with_topic_ids, subtopic_ids)

    @staticmethod
    def list_models(db_session):
        def instances_with_at_least_one_question(model):
            return db_session.query(model). \
                join(Question). \
                group_by(model). \
                having(func.count(Question.id) > 0). \
                all()
        return {
            u'autor': instances_with_at_least_one_question(Author),
            u'informe': instances_with_at_least_one_question(Report),
            u'área de gestión': instances_with_at_least_one_question(SubTopic),
            u'ministerio': instances_with_at_least_one_question(Topic)
        }

    @staticmethod
    def get_question(question_id):
        question = Question.query.get(question_id)
        return question

    def search_from_url(self):
        query = self.query_from_url()
        return self.search(query)

    def delete_results_from_url(self, db_session):
        query = self.query_from_url()
        results = self._search_questions(query)
        for question, keywords in results:
            db_session.delete(question)
        db_session.commit()
        self.restart_text_classifier()
        return

    def search(self, query):
        results = self._search_questions(query)
        return self._paginate(results, query)

    def _search_questions(self, query):
        if query['text'] is not None:
            g.similarity_cutoff = 1.1
            results = self._search_similar(query)
        else:
            results = Question.query.all()
            results = self._order_results(results, query)
            results = [(result, []) for result in results]
        results = self._filter_results(results, query['filters'])
        return results

    @staticmethod
    def _order_results(results, query):
        if query['order'] in ('asc', 'desc'):
            return sorted(results, key=lambda x: (x.report.name, x.number),
                          reverse=query['order'] == 'desc')
        elif query['order'] in ('date-asc', 'date-desc'):
            none_res = filter(lambda x: x.question_date is None, results)
            not_none_res = filter(lambda x: x.question_date is not None, results)
            ord_results = sorted(not_none_res, key=lambda x: x.question_date,
                                 reverse=query['order'] == 'date-desc')
            return list(ord_results) + list(none_res)
        else:
            return results

    def _paginate(self, results, query):
        per_page = 'por-pagina' in query and int(query['por-pagina']) or self.per_page
        pagination = {
            'current_page': query['current_page'],
            'total_pages': int(math.ceil(len(results) / float(per_page))),
            'total_results': len(results)
        }
        from_position = (pagination['current_page'] - 1) * per_page
        to_position = pagination['current_page'] * per_page
        return {
            'pagination': pagination,
            'result_list': results[from_position:to_position],
            'query': query
        }

    @staticmethod
    def _pass_filter(result, filters):
        """Recives an item of the results list [(result, best_words)]
            and a dict of filter_ids and decides whether that element is
            accepted by the filter or not.
        """
        result_only = result[0]
        comparisions = []
        for filter_attr, filter_value in filters.items():
            if filter_value['filter_value'] and len(filter_value['filter_value']) > 0:
                compare_to = filter_value['filter_value'][0].id
            else:
                compare_to = filter_value['filter_value']
            if filter_value['filter_by'] == 'igualdad':
                comparisions.append(getattr(result_only, filter_attr) == compare_to)
            else:
                comparisions.append(getattr(result_only, filter_attr) != compare_to)
        return all(comparisions)

    @staticmethod
    def _collect_filter_values(filters):
        filter_models = {
            'ministerio': ('topic_id', Topic),
            'area': ('subtopic_id', SubTopic),
            'autor': ('author_id', Author),
            'informe': ('report_id', Report)
        }
        filter_values = {}
        for filter_name, filter_model in filter_models.items():
            if filter_name in filters.keys():
                comparision_key = filter_name + '-comparacion'
                filter_info = {
                    'filter_by': comparision_key in filters and filters[comparision_key] or 'igualdad',
                    'filter_value': None
                }
                if len(filters[filter_name]) > 0:
                    filter_info['filter_value'] = filter_model[1].query.filter_by(name=filters[filter_name]).all()
                filter_values[filter_model[0]] = filter_info
        return filter_values

    def _filter_results(self, results, filters):
        filter_values = self._collect_filter_values(filters)
        filtered_questions = filter(lambda result: self._pass_filter(result, filter_values), results)
        if 'creado-en' in filters:
            created_at = datetime.strptime(filters['creado-en'], '%Y-%m-%d %H:%M:%S')
            filtered_questions = filter(lambda x: x[0].created_at == created_at, filtered_questions)
        return list(filtered_questions)

    def get_similar_to(self, question):
        query = self.query_from_url()
        query['id'] = question.id
        if query['based_on'] == 'pregunta':
            query['text'] = 'q' + str(question.id)
        elif query['based_on'] == 'respuesta':
            query['text'] = 'r' + str(question.id)
        else:
            query['text'] = question.context + ' ' + question.body + ' ' + question.answer
        return self.search(query)

    def _search_similar(self, query):
        question_id = query['text']
        all_questions = Question.query.all()
        if self.text_classifier is None:
            return []
        if query['target'] == 'preguntas':
            id_list = ['q' + str(q.id) for q in all_questions
                       if q.body is not None and len(q.body) > 0]
        elif query['target'] == 'respuestas':
            id_list = ['r' + str(q.id) for q in all_questions
                       if q.answer is not None and len(q.answer) > 0]
        else:
            id_list = ['q' + str(q.id) for q in all_questions
                       if q.body is not None and len(q.body) > 0] + ['r' + str(q.id) for q in all_questions
                       if q.answer is not None and len(q.answer) > 0]
        if not isinstance(question_id, str):
            question_id = str(question_id)
        per_page = 'por-pagina' in query and int(query['por-pagina']) or self.per_page
        if id_list:
            max_options = min(per_page, len(id_list))
        else:
            max_options = per_page
        ids_sim, dist, best_words = self.text_classifier.get_similar(
            question_id, max_similars=max_options, filter_list=id_list, term_diff_max_rank=40)
        ids_sim = self._clean_ids(ids_sim, query)
        results = []
        for qid in ids_sim:
            results.append(Question.query.get(qid))
        return zip(results, best_words, dist)

    def suggest_tags(self, tag_type, question_id):
        question = Question.query.get(question_id)
        if tag_type == 'topics':
            tags, vals = self.text_classifier.classify(
                'topics', ['q' + str(question_id)])
            model = Topic
        else:
            classifier_name = str(question.topic_id) + "_" + tag_type
            model = SubTopic

            if classifier_name in dir(self.text_classifier):
                tags, vals = self.text_classifier.classify(
                    classifier_name, ['q' + str(question_id)])
            elif question.topic is not None:
                subtopics = question.topic.subtopics
                return list(sorted([x.name for x in subtopics]))
            else:
                subtopics = SubTopic.query.all()
                return list(sorted([s.name for s in subtopics]))

        tag_names = [model.query.get(idt).name for idt in tags]
        sorted_tags = [x for (y, x) in
                       sorted(zip(vals.tolist()[0], tag_names))]
        return list(reversed(sorted_tags))

    @staticmethod
    def query_from_url():
        filter_titles = [
            'ministerio', 'ministerio-comparacion',
            'area', 'area-comparacion',
            'autor', 'autor-comparacion',
            'informe', 'informe-comparacion',
            'organismo-requerido',
            'pregunta', 'creado-en'
        ]
        query = {
            'text': request.args.get('q'),
            'target': request.args.get('buscar-dentro-de', ''),
            'based_on': request.args.get('buscar-usando', ''),
            'current_page': int(request.args.get('pagina', 1)),
            'can_add_more_filters': True,
            'order': request.args.get('orden', 'asc'),
            'filters': {t: request.args.get(t).lower() for t in filter_titles
                        if request.args.get(t) is not None},
        }
        if request.args.get('por-pagina'):
            query['por-pagina'] = request.args.get('por-pagina')
        return query

    @staticmethod
    def _clean_ids(ids, query):
        ids = map(lambda x: x[1:], ids)
        if 'id' in query:
            question_id = str(query['id'])
            ids = filter(lambda x: x != question_id, ids)
        seen = set()
        seen_add = seen.add
        return [int(x) for x in ids if not (x in seen or seen_add(x))]

    @staticmethod
    def url_maker(query, page=None):
        args = {}
        if 'text' in query and query['text'] is not None:
            args['q'] = query['text']
        if 'order' in query:
            args['order'] = query['order']
        for title, value in query['filters'].items():
            args[title] = value
        if page is not None:
            args['pagina'] = page
        return url_for('search', **args)
